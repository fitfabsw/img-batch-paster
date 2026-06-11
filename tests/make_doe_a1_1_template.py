"""產生 AI DOE-A1-1(延伸) 測試用 Excel 範本，欄位與範例一致。

  Title | C | D | BC | E | CCD | G | H | TC | I | K | TE
  下方為留空、加框的資料列（app 會自動偵測表格、把 2EG/2JG 寫入 Title 欄並橫向貼圖）。

跑法：.venv/bin/python tests/make_doe_a1_1_template.py
輸出：tests/fixtures/doe-a1-1-template.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "fixtures" / "doe-a1-1-template.xlsx"

HEADERS = ["Title", "C", "D", "BC", "E", "CCD", "G", "H", "TC", "I", "K", "TE"]
DATA_ROWS = 4   # 留空資料列（2EG / 2JG + 緩衝）

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
ncols = len(HEADERS)

# 表頭列（粗體、置中、加框）
for ci, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True)
    c.alignment = center
    c.border = border

# 資料列：全部留空、加框（供偵測表格 + 貼圖）
for r in range(2, 2 + DATA_ROWS):
    for ci in range(1, ncols + 1):
        ws.cell(row=r, column=ci).border = border
    ws.row_dimensions[r].height = 85   # 影像列高

# 欄寬：Title 窄、影像欄寬
ws.column_dimensions["A"].width = 10
for ci in range(2, ncols + 1):
    ws.column_dimensions[get_column_letter(ci)].width = 16
ws.row_dimensions[1].height = 22

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"wrote template: {OUT}  ({ncols} cols x {1 + DATA_ROWS} rows)")
