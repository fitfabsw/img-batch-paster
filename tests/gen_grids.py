"""把 sample_xlsx/templates/*.xlsx 轉成前端 grid JSON（cells/borders/cols），
供 test_placement.mjs 餵給真實的前端 placement 函式。格式比照 app.py 的 api_template_excel_grid。

用法：.venv/bin/python tests/gen_grids.py  →  寫出 tests/fixtures/sample_xlsx/grids.json
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

tpl_dir = Path(__file__).parent / "fixtures" / "sample_xlsx" / "templates"
out = Path(__file__).parent / "fixtures" / "sample_xlsx" / "grids.json"

def grid_of(path: Path) -> dict:
    ws = load_workbook(str(path)).active
    max_col = max(ws.max_column or 1, 12)
    max_row = max(ws.max_row or 1, 20)
    cols = [{"letter": get_column_letter(c)} for c in range(1, max_col + 4)]
    cells, borders = [], []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cells.append({"r": cell.row, "c": cell.column, "text": str(cell.value)})
            b = cell.border
            if any(bool(s and s.style) for s in (b.top, b.right, b.bottom, b.left)):
                borders.append({"r": cell.row, "c": cell.column})
    return {"cols": cols, "cells": cells, "borders": borders}

import re
grids = {p.stem: grid_of(p) for p in sorted(tpl_dir.glob("*.xlsx"))
         if re.match(r"[hv][1-4]_", p.stem)}   # 只取 8 個 canonical fixtures，忽略臨時檔
out.write_text(json.dumps(grids, ensure_ascii=False, indent=0))
print(f"wrote {out} ({len(grids)} grids)")
