"""產生 xlsx 依檔名 測試範例：4 張圖 + 6 個範本（横式/直式 × 全空/index/index+group）。

用法：.venv/bin/python tests/fixtures/sample_xlsx/make_samples.py
詳見同目錄 README.md。
"""
from pathlib import Path
from PIL import Image, ImageDraw
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

base = Path(__file__).parent
img_dir = base / "images"
tpl_dir = base / "templates"
img_dir.mkdir(parents=True, exist_ok=True)
tpl_dir.mkdir(parents=True, exist_ok=True)

# 4 張測試圖（檔名 = {group}-{index}）
colors = {"AAA-1": (220, 120, 120), "AAA-2": (120, 180, 120),
          "BBB-2": (120, 140, 210), "BBB-3": (210, 180, 110)}
for name, rgb in colors.items():
    im = Image.new("RGB", (240, 180), rgb)
    ImageDraw.Draw(im).text((90, 80), name, fill="white")
    im.save(img_dir / f"{name}.png")

thin = Side(style="thin", color="000000")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")

def grid(ws, r0, c0, nrows, ncols):
    for r in range(r0, r0 + nrows):
        ws.row_dimensions[r].height = 60
        for c in range(c0, c0 + ncols):
            ws.cell(r, c).border = BORD
            ws.cell(r, c).alignment = CEN
            ws.column_dimensions[ws.cell(r, c).column_letter].width = 14

def put(ws, r, c, v):
    ws.cell(r, c, v)
    ws.cell(r, c).font = Font(bold=True)

def save(wb, name):
    wb.save(tpl_dir / name)

INDEX = ["2", "3", "4"]      # 3 個 index
GROUP = ["BBB", "CCC"]       # 2 個 group（含一個對不到的 CCC）

# h/v 同號＝同邏輯：*1=全空 *2=只有 index *3=index+group *4=只有 group
# 横式：Group 在左欄、Index 在頂列
wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 3, 4); save(wb, "h1_empty.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 4, 4)
for j, v in enumerate(INDEX): put(ws, 2, 3 + j, v)
save(wb, "h2_index.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 1 + len(GROUP), 1 + len(INDEX))
put(ws, 2, 2, "Title")
for j, v in enumerate(INDEX): put(ws, 2, 3 + j, v)
for i, v in enumerate(GROUP): put(ws, 3 + i, 2, v)
save(wb, "h3_index_group.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 1 + len(GROUP), 4)
for i, v in enumerate(GROUP): put(ws, 3 + i, 2, v)
save(wb, "h4_group.xlsx")

# 直式：Group 在頂列、Index 在左欄
wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 4, 4); save(wb, "v1_empty.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 1 + len(INDEX), 4)
for i, v in enumerate(INDEX): put(ws, 3 + i, 2, v)
save(wb, "v2_index.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 1 + len(INDEX), 1 + len(GROUP))
put(ws, 2, 2, "Title")
for j, v in enumerate(GROUP): put(ws, 2, 3 + j, v)
for i, v in enumerate(INDEX): put(ws, 3 + i, 2, v)
save(wb, "v3_index_group.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; grid(ws, 2, 2, 4, 1 + len(GROUP))
for j, v in enumerate(GROUP): put(ws, 2, 3 + j, v)
save(wb, "v4_group.xlsx")

print("done →", tpl_dir)
