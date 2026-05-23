"""Excel batch image writer (openpyxl)."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from PIL import Image as PILImage
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, TwoCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import pixels_to_EMU


@dataclass
class CellPlacement:
    path: Path | None       # None when this is a text-only placement
    row: int                # 1-based
    col: int                # 1-based
    span_cols: int = 1
    span_rows: int = 1
    text: str | None = None
    font_pt: float = 12.0


# Excel 像素換算
# column: px ≈ width * MDW + 5   MDW (Max Digit Width) 隨字體變
# row:    px ≈ height * 4/3      (H=15 → 20px)
def excel_col_to_px(w: float, mdw: float = 7.0) -> float:
    return w * mdw + 5

def excel_row_to_px(h: float) -> float:
    return h * 4 / 3


# contain 模式四邊留白比例 (每邊佔 cell 的 5%)
_CONTAIN_INSET = 0.05


def _default_mdw(wb) -> float:
    """根據範本預設字體大小估算 MDW。
    Calibri/新細明體 11 → 7；12 → ~8；14 → ~9。粗略線性。
    """
    try:
        sz = float(wb._fonts[0].sz or 11)
    except Exception:
        sz = 11.0
    if sz <= 11:
        return 7.0
    # 每多 1 pt 字大約多 ~0.7 px MDW
    return 7.0 + (sz - 11) * 0.85


def _col_width(ws, col_idx: int) -> float:
    """範本可能用 <col min=A max=B width=W> 範圍宣告；要走全部 column_dimensions 才能正確取到。"""
    for cd in ws.column_dimensions.values():
        if cd.min is None or cd.max is None:
            continue
        if cd.min <= col_idx <= cd.max and cd.width is not None:
            return cd.width
    return ws.sheet_format.defaultColWidth or 8.43


def _row_height(ws, row_idx: int) -> float:
    rd = ws.row_dimensions.get(row_idx)
    if rd is not None and rd.height is not None:
        return rd.height
    return ws.sheet_format.defaultRowHeight or 15.0


def _cell_pixel_size(ws, col: int, row: int, mdw: float = 7.0) -> tuple[float, float]:
    return excel_col_to_px(_col_width(ws, col), mdw), excel_row_to_px(_row_height(ws, row))


def _placement_pixel_size(ws, p: CellPlacement, mdw: float = 7.0) -> tuple[float, float]:
    w = sum(_cell_pixel_size(ws, p.col + i, p.row, mdw)[0] for i in range(p.span_cols))
    h = sum(_cell_pixel_size(ws, p.col, p.row + j, mdw)[1] for j in range(p.span_rows))
    return w, h


def _stretch_resize(img_path: Path, target_w: float, target_h: float, tmp_dir: Path) -> Path:
    """非等比拉伸到 target 大小。"""
    target_w, target_h = max(1, int(round(target_w))), max(1, int(round(target_h)))
    with PILImage.open(img_path) as im:
        if im.mode == "P":
            im = im.convert("RGB")
        im = im.resize((target_w, target_h), PILImage.LANCZOS)
        tmp_dir.mkdir(parents=True, exist_ok=True)
        out = tmp_dir / f"_fill_{img_path.stem}_{target_w}x{target_h}{img_path.suffix or '.png'}"
        save_im = im if (out.suffix.lower() == ".png" or im.mode != "RGBA") else im.convert("RGB")
        save_im.save(out)
        return out


def _cover_crop(img_path: Path, target_w: float, target_h: float, tmp_dir: Path) -> Path:
    """等比放大圖片直到至少有一邊貼齊 cell 大小，超出的部分 center-crop 掉。"""
    target_w, target_h = max(1, int(round(target_w))), max(1, int(round(target_h)))
    with PILImage.open(img_path) as im:
        if im.mode == "P":
            im = im.convert("RGBA" if "transparency" in im.info else "RGB")
        iw, ih = im.size
        scale = max(target_w / iw, target_h / ih)
        new_w = max(target_w, int(round(iw * scale)))
        new_h = max(target_h, int(round(ih * scale)))
        im = im.resize((new_w, new_h), PILImage.LANCZOS)
        left = (new_w - target_w) // 2
        top = (new_h - target_h) // 2
        im = im.crop((left, top, left + target_w, top + target_h))
        tmp_dir.mkdir(parents=True, exist_ok=True)
        out = tmp_dir / f"_crop_{img_path.stem}_{target_w}x{target_h}{img_path.suffix or '.png'}"
        # 確保 PIL 支援的格式
        save_kwargs = {}
        if im.mode == "RGBA" and out.suffix.lower() not in (".png",):
            im = im.convert("RGB")
        im.save(out, **save_kwargs)
        return out


def write_xlsx(
    placements: list[CellPlacement],
    out_path: Path,
    template: Path | None = None,
    sheet_name: str | None = None,
    embed_in_cell: bool = False,
    lock_images: bool = True,
    img_fit: str = "cover",
) -> Path:
    """Embed images either as floating drawings (default) or as cell content (DISPIMG).

    lock_images=True (default) adds picLocks so floating images cannot be moved
    / resized / selected, mimicking cell-embedded behavior visually.
    """
    if embed_in_cell:
        return _write_xlsx_in_cell(placements, out_path, template, sheet_name, img_fit=img_fit)

    if template and Path(template).is_file():
        wb = load_workbook(str(template))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active

    mdw = _default_mdw(wb)

    for p in placements:
        cell = ws.cell(row=p.row, column=p.col)
        if p.text is not None:
            cell.value = p.text
            if p.font_pt:
                cell.font = cell.font.copy(size=p.font_pt, bold=True)
            continue

        if p.path is None or not Path(p.path).is_file():
            continue

        target_w, target_h = _placement_pixel_size(ws, p, mdw)

        if img_fit == "fill":
            # TwoCellAnchor 拉伸到 cell 範圍 (會變形)
            img = XLImage(str(p.path))
            img.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=p.col - 1, colOff=0, row=p.row - 1, rowOff=0),
                to=AnchorMarker(col=p.col - 1 + p.span_cols, colOff=0,
                                row=p.row - 1 + p.span_rows, rowOff=0),
                editAs="oneCell",
            )
            ws._images.append(img)
        elif img_fit == "contain":
            # 保長寬比、放入 cell 並四邊留白置中
            with PILImage.open(p.path) as im:
                iw, ih = im.size
            aspect = ih / iw if iw else 1.0
            avail_w = target_w * (1 - 2 * _CONTAIN_INSET)
            avail_h = target_h * (1 - 2 * _CONTAIN_INSET)
            fit_w = avail_w
            fit_h = fit_w * aspect
            if fit_h > avail_h and avail_h > 0:
                fit_h = avail_h
                fit_w = fit_h / aspect if aspect else avail_w
            col_off_px = max(0.0, (target_w - fit_w) / 2)
            row_off_px = max(0.0, (target_h - fit_h) / 2)
            img = XLImage(str(p.path))
            img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=p.col - 1, colOff=int(round(pixels_to_EMU(col_off_px))),
                    row=p.row - 1, rowOff=int(round(pixels_to_EMU(row_off_px))),
                ),
                ext=XDRPositiveSize2D(
                    cx=int(round(pixels_to_EMU(fit_w))),
                    cy=int(round(pixels_to_EMU(fit_h))),
                ),
            )
            ws._images.append(img)
        else:
            # cover (default)：保長寬比、center-crop、剛好填滿 cell
            cropped = _cover_crop(Path(p.path), target_w, target_h, out_path.parent / "_crops")
            img = XLImage(str(cropped))
            img.width = int(round(target_w))
            img.height = int(round(target_h))
            ws.add_image(img, f"{get_column_letter(p.col)}{p.row}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    if lock_images:
        _lock_drawing_images(out_path)
    return out_path


def _lock_drawing_images(xlsx_path: Path) -> None:
    """Inject a:picLocks (noMove + noResize + noSelect) into every drawing's
    cNvPicPr so images become non-interactive once embedded."""
    import zipfile
    import shutil
    import tempfile

    picLocks_xml = (
        '<a:picLocks xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'noChangeAspect="1" noMove="1" noResize="1" noSelect="1"/>'
    )

    tmp = Path(tempfile.mkdtemp())
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            zf.extractall(tmp)
        drawings_dir = tmp / "xl" / "drawings"
        if not drawings_dir.is_dir():
            return
        changed = False
        for d in drawings_dir.glob("drawing*.xml"):
            content = d.read_text(encoding="utf-8")
            # 自封閉 <xdr:cNvPicPr/>
            new = content.replace(
                "<xdr:cNvPicPr/>",
                f"<xdr:cNvPicPr>{picLocks_xml}</xdr:cNvPicPr>",
            )
            # 有屬性但無內容 <xdr:cNvPicPr ... />
            import re
            new = re.sub(
                r"<xdr:cNvPicPr([^>]*?)/>",
                lambda m: f"<xdr:cNvPicPr{m.group(1)}>{picLocks_xml}</xdr:cNvPicPr>",
                new,
            )
            if new != content:
                d.write_text(new, encoding="utf-8")
                changed = True
        if not changed:
            return
        xlsx_path.unlink()
        with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in sorted(tmp.rglob("*")):
                if f.is_file():
                    zf.write(f, f.relative_to(tmp))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _write_xlsx_in_cell(
    placements: list[CellPlacement],
    out_path: Path,
    template: Path | None,
    sheet_name: str | None,
    img_fit: str = "contain",
) -> Path:
    """Embed images as cell content using Microsoft Excel 365 RichValue schema.

    Strategy:
    1. openpyxl writes the workbook with text-only cells; image-target cells
       are left empty for us to inject.
    2. Post-process the saved zip to add: xl/metadata.xml,
       xl/richData/{rdRichValues,rdRichValueStructure,rdRichValueTypes,
       richValueRel}.xml + relationships, plus updated [Content_Types].xml
       and xl/_rels/workbook.xml.rels.
    3. Insert <c r="X" t="e" vm="N"><v>#VALUE!</v></c> into the worksheet's
       sheetData at each image cell position; vm references valueMetadata.
    """
    import zipfile
    import shutil
    import tempfile
    from xml.etree import ElementTree as ET

    if template and Path(template).is_file():
        wb = load_workbook(str(template))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active

    image_records: list[tuple[Path, str]] = []          # (src_path, media_name)
    image_cells: list[tuple[str, int, int, int]] = []   # (col_letter, row, rv_index_0based, vm_1based)
    mdw = _default_mdw(wb)

    for p in placements:
        if p.text is not None:
            cell = ws.cell(row=p.row, column=p.col)
            cell.value = p.text
            if p.font_pt:
                cell.font = cell.font.copy(size=p.font_pt, bold=True)
            continue
        if p.path is None or not Path(p.path).is_file():
            continue
        # 依 img_fit 預處理圖片：Excel embed 模式預設 contain，所以 cover/fill 都要 PIL 先處理
        src_path = Path(p.path)
        if img_fit in ("cover", "fill"):
            tw, th = _placement_pixel_size(ws, p, mdw)
            tmp_dir = out_path.parent / "_crops"
            src_path = (_cover_crop if img_fit == "cover" else _stretch_resize)(
                src_path, tw, th, tmp_dir
            )
        ext = src_path.suffix.lower() or ".png"
        media_name = f"image_rv{len(image_records) + 1}{ext}"
        image_records.append((src_path, media_name))
        image_cells.append((
            get_column_letter(p.col), p.row,
            len(image_records) - 1,
            len(image_cells) + 1,
        ))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    if not image_cells:
        return out_path

    sheet_name_active = ws.title
    _inject_richvalue(out_path, image_records, image_cells, sheet_name_active)
    return out_path


# 命名空間常數
_NS_SS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
_NS_RD = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"
_NS_RD2 = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata2"
_NS_RVR = "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
_NS_MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"

# 各種 relationship type
_REL_SHEET_METADATA = f"{_NS_R}/sheetMetadata"
_REL_RD_RV = "http://schemas.microsoft.com/office/2017/06/relationships/rdRichValue"
_REL_RD_RVS = "http://schemas.microsoft.com/office/2017/06/relationships/rdRichValueStructure"
_REL_RVT = "http://schemas.microsoft.com/office/2022/10/relationships/richValueTypes"
_REL_RVR = "http://schemas.microsoft.com/office/2022/10/relationships/richValueRel"

# futureMetadata 內 ext uri (real MS Excel value)
_RV_EXT_URI = "{3e2802c4-a4d2-4d8b-9148-e3be6c30e623}"


def _inject_richvalue(
    xlsx_path: Path,
    image_records: list[tuple[Path, str]],
    image_cells: list[tuple[str, int, int, int]],
    sheet_name_active: str,
) -> None:
    import zipfile
    import shutil
    import tempfile
    import re
    from xml.etree import ElementTree as ET

    tmp = Path(tempfile.mkdtemp())
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            zf.extractall(tmp)

        # 1. 複製圖片
        media_dir = tmp / "xl" / "media"
        media_dir.mkdir(parents=True, exist_ok=True)
        for src, name in image_records:
            shutil.copy2(src, media_dir / name)

        rd_dir = tmp / "xl" / "richData"
        (rd_dir / "_rels").mkdir(parents=True, exist_ok=True)

        # 2. rdRichValues.xml
        rv_xml = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>']
        rv_xml.append(f'<rvData xmlns="{_NS_RD}" count="{len(image_records)}">')
        for i in range(len(image_records)):
            rv_xml.append(f'<rv s="0"><v>{i}</v><v>5</v></rv>')
        rv_xml.append('</rvData>')
        (rd_dir / "rdrichvalue.xml").write_text("".join(rv_xml), encoding="utf-8")

        # 3. rdrichvaluestructure.xml
        rvs_xml = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<rvStructures xmlns="{_NS_RD}" count="1">'
            f'<s t="_localImage">'
            f'<k n="_rvRel:LocalImageIdentifier" t="i"/>'
            f'<k n="CalcOrigin" t="i"/>'
            f'</s>'
            f'</rvStructures>'
        )
        (rd_dir / "rdrichvaluestructure.xml").write_text(rvs_xml, encoding="utf-8")

        # 4. rdRichValueTypes.xml
        rvt_xml = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<rvTypesInfo xmlns="{_NS_RD2}" xmlns:mc="{_NS_MC}" mc:Ignorable="x">'
            f'<global>'
            f'<keyFlags>'
            f'<key name="_Self"><flag name="ExcludeFromFile" value="1"/><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_DisplayString"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_Flags"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_Format"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_SubLabel"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_Attribution"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_Icon"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_Display"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_CanonicalPropertyNames"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'<key name="_ClassificationId"><flag name="ExcludeFromCalcComparison" value="1"/></key>'
            f'</keyFlags>'
            f'</global>'
            f'</rvTypesInfo>'
        )
        (rd_dir / "rdRichValueTypes.xml").write_text(rvt_xml, encoding="utf-8")

        # 5. richValueRel.xml
        rvr_xml = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>']
        rvr_xml.append(f'<richValueRels xmlns="{_NS_RVR}" xmlns:r="{_NS_R}">')
        for i in range(len(image_records)):
            rvr_xml.append(f'<rel r:id="rId{i + 1}"/>')
        rvr_xml.append('</richValueRels>')
        (rd_dir / "richValueRel.xml").write_text("".join(rvr_xml), encoding="utf-8")

        # 6. richValueRel.xml.rels — 把 rId 對應到 image media
        rvr_rels = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>']
        rvr_rels.append(f'<Relationships xmlns="{_NS_PKG}">')
        for i, (_src, name) in enumerate(image_records):
            rvr_rels.append(
                f'<Relationship Id="rId{i + 1}" Type="{_NS_R}/image" Target="../media/{name}"/>'
            )
        rvr_rels.append('</Relationships>')
        (rd_dir / "_rels" / "richValueRel.xml.rels").write_text("".join(rvr_rels), encoding="utf-8")

        # 7. metadata.xml
        md = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>']
        md.append(
            f'<metadata xmlns="{_NS_SS}" xmlns:xlrd="{_NS_RD}">'
        )
        md.append('<metadataTypes count="1">')
        md.append(
            '<metadataType name="XLRICHVALUE" minSupportedVersion="120000" '
            'copy="1" pasteAll="1" pasteValues="1" merge="1" splitFirst="1" '
            'rowColShift="1" clearFormats="1" clearComments="1" assign="1" '
            'coerce="1"/>'
        )
        md.append('</metadataTypes>')
        md.append(f'<futureMetadata name="XLRICHVALUE" count="{len(image_cells)}">')
        for _col, _row, rv_idx, _vm in image_cells:
            md.append(
                f'<bk><extLst>'
                f'<ext uri="{_RV_EXT_URI}"><xlrd:rvb i="{rv_idx}"/></ext>'
                f'</extLst></bk>'
            )
        md.append('</futureMetadata>')
        md.append(f'<valueMetadata count="{len(image_cells)}">')
        for i in range(len(image_cells)):
            md.append(f'<bk><rc t="1" v="{i}"/></bk>')
        md.append('</valueMetadata>')
        md.append('</metadata>')
        (tmp / "xl" / "metadata.xml").write_text("".join(md), encoding="utf-8")

        # 8. 修改 worksheet — 注入 image cells
        # 找出 active sheet 對應的 xml 檔
        wb_xml_path = tmp / "xl" / "workbook.xml"
        ET.register_namespace("", _NS_SS)
        ET.register_namespace("r", _NS_R)
        wb_tree = ET.parse(wb_xml_path)
        sheets_el = wb_tree.getroot().find(f"{{{_NS_SS}}}sheets")
        sheet_rel_id = None
        for s in sheets_el.findall(f"{{{_NS_SS}}}sheet"):
            if s.get("name") == sheet_name_active:
                sheet_rel_id = s.get(f"{{{_NS_R}}}id")
                break
        wb_rels_path = tmp / "xl" / "_rels" / "workbook.xml.rels"
        ws_target = None
        if sheet_rel_id:
            wbr_tree = ET.parse(wb_rels_path)
            for r in wbr_tree.getroot().findall(f"{{{_NS_PKG}}}Relationship"):
                if r.get("Id") == sheet_rel_id:
                    ws_target = r.get("Target")
                    break
        if ws_target is None:
            sheets = sorted((tmp / "xl" / "worksheets").glob("sheet*.xml"))
            ws_xml_path = sheets[0] if sheets else None
        else:
            # Target 可能是絕對(/xl/...) 或相對(workbook.xml 位置, 即 xl/)
            if ws_target.startswith("/"):
                ws_xml_path = (tmp / ws_target.lstrip("/")).resolve()
            else:
                ws_xml_path = (tmp / "xl" / ws_target).resolve()

        if ws_xml_path is None or not ws_xml_path.exists():
            raise RuntimeError(f"找不到 worksheet XML: target={ws_target}")

        ws_tree = ET.parse(ws_xml_path)
        ws_root = ws_tree.getroot()
        sd = ws_root.find(f"{{{_NS_SS}}}sheetData")

        # row map: row_num -> row element
        rows_by_num = {}
        for r in sd.findall(f"{{{_NS_SS}}}row"):
            try:
                rows_by_num[int(r.get("r"))] = r
            except (TypeError, ValueError):
                continue

        def _col_idx(letter: str) -> int:
            return column_index_from_string(letter)

        for col_letter, row_num, _rv_idx, vm_idx in image_cells:
            row_el = rows_by_num.get(row_num)
            if row_el is None:
                row_el = ET.SubElement(sd, f"{{{_NS_SS}}}row")
                row_el.set("r", str(row_num))
                rows_by_num[row_num] = row_el
            cell_ref = f"{col_letter}{row_num}"
            existing_c = None
            for c in row_el.findall(f"{{{_NS_SS}}}c"):
                if c.get("r") == cell_ref:
                    existing_c = c
                    break
            if existing_c is not None:
                # 保留範本既有的 style (s 屬性)，覆寫 type/vm/value
                s_attr = existing_c.get("s")
                for child in list(existing_c):
                    existing_c.remove(child)
                existing_c.attrib.clear()
                existing_c.set("r", cell_ref)
                if s_attr:
                    existing_c.set("s", s_attr)
                existing_c.set("t", "e")
                existing_c.set("vm", str(vm_idx))
                v = ET.SubElement(existing_c, f"{{{_NS_SS}}}v")
                v.text = "#VALUE!"
            else:
                c_el = ET.SubElement(row_el, f"{{{_NS_SS}}}c")
                c_el.set("r", cell_ref)
                c_el.set("t", "e")
                c_el.set("vm", str(vm_idx))
                v_el = ET.SubElement(c_el, f"{{{_NS_SS}}}v")
                v_el.text = "#VALUE!"

        # 排序 cells in each row by column index
        for r in sd.findall(f"{{{_NS_SS}}}row"):
            cs = list(r.findall(f"{{{_NS_SS}}}c"))
            cs.sort(key=lambda c: _col_idx(re.match(r"[A-Z]+", c.get("r") or "A1").group()))
            for c in cs:
                r.remove(c)
            for c in cs:
                r.append(c)

        # 排序 rows by row number
        rows = list(sd.findall(f"{{{_NS_SS}}}row"))
        rows.sort(key=lambda r: int(r.get("r") or 0))
        for r in rows:
            sd.remove(r)
        for r in rows:
            sd.append(r)

        ws_tree.write(ws_xml_path, xml_declaration=True, encoding="UTF-8")

        # 9. [Content_Types].xml
        ct_path = tmp / "[Content_Types].xml"
        ET.register_namespace("", _NS_CT)
        ct_tree = ET.parse(ct_path)
        ct_root = ct_tree.getroot()
        existing_defaults = {d.get("Extension") for d in ct_root.findall(f"{{{_NS_CT}}}Default")}
        for ext, mime in [("png", "image/png"), ("jpg", "image/jpeg"), ("jpeg", "image/jpeg")]:
            if ext not in existing_defaults:
                el = ET.SubElement(ct_root, f"{{{_NS_CT}}}Default")
                el.set("Extension", ext)
                el.set("ContentType", mime)
        overrides = [
            ("/xl/metadata.xml", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheetMetadata+xml"),
            ("/xl/richData/rdrichvalue.xml", "application/vnd.ms-excel.rdrichvalue+xml"),
            ("/xl/richData/rdrichvaluestructure.xml", "application/vnd.ms-excel.rdrichvaluestructure+xml"),
            ("/xl/richData/rdRichValueTypes.xml", "application/vnd.ms-excel.rdrichvaluetypes+xml"),
            ("/xl/richData/richValueRel.xml", "application/vnd.ms-excel.richvaluerel+xml"),
        ]
        existing_overrides = {o.get("PartName") for o in ct_root.findall(f"{{{_NS_CT}}}Override")}
        for part, mime in overrides:
            if part in existing_overrides:
                continue
            el = ET.SubElement(ct_root, f"{{{_NS_CT}}}Override")
            el.set("PartName", part)
            el.set("ContentType", mime)
        ct_tree.write(ct_path, xml_declaration=True, encoding="UTF-8")

        # 10. xl/_rels/workbook.xml.rels
        ET.register_namespace("", _NS_PKG)
        wbr_tree = ET.parse(wb_rels_path)
        wbr_root = wbr_tree.getroot()
        used_ids = {r.get("Id") for r in wbr_root.findall(f"{{{_NS_PKG}}}Relationship")}

        def _next_id(prefix: str) -> str:
            i = 1
            while f"{prefix}{i}" in used_ids:
                i += 1
            new_id = f"{prefix}{i}"
            used_ids.add(new_id)
            return new_id

        rels_to_add = [
            (_REL_SHEET_METADATA, "metadata.xml"),
            (_REL_RD_RV, "richData/rdrichvalue.xml"),
            (_REL_RD_RVS, "richData/rdrichvaluestructure.xml"),
            (_REL_RVT, "richData/rdRichValueTypes.xml"),
            (_REL_RVR, "richData/richValueRel.xml"),
        ]
        existing_rel_pairs = {
            (r.get("Type"), r.get("Target"))
            for r in wbr_root.findall(f"{{{_NS_PKG}}}Relationship")
        }
        for rtype, target in rels_to_add:
            if (rtype, target) in existing_rel_pairs:
                continue
            el = ET.SubElement(wbr_root, f"{{{_NS_PKG}}}Relationship")
            el.set("Id", _next_id("rIdRV"))
            el.set("Type", rtype)
            el.set("Target", target)
        wbr_tree.write(wb_rels_path, xml_declaration=True, encoding="UTF-8")

        # 11. 重新打包
        xlsx_path.unlink()
        with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in sorted(tmp.rglob("*")):
                if f.is_file():
                    zf.write(f, f.relative_to(tmp))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def parse_cell(cell_ref: str) -> tuple[int, int]:
    """'B5' → (col=2, row=5)"""
    import re
    m = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref.strip())
    if not m:
        raise ValueError(f"無效 cell 表示: {cell_ref}")
    return column_index_from_string(m.group(1).upper()), int(m.group(2))
